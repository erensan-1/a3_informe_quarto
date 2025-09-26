import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Configurar tipografía moderna
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['font.size'] = 10

# Cargar Excel
file_path = "a3_asignaciones_vc.xlsx"
df = pd.read_excel(file_path)

# Diccionario de nombres amigables de estados
estado_labels = {
    "A3_AsignadaVia": "Asignada Vía",
    "A3_AsignadaViaAlternativa": "Asignada Vía Alternativa",
    "A3_AsignadaUnidadPoblacional": "Asignada Unidad Poblacional",
    "A3_AsignarViaNueva": "Asignar Vía Nueva",
    "A3_Aislada": "Aislada"
}

# Diccionario de abreviaturas SOLO para el gráfico
estado_abrev = {
    "Asignada Vía Alternativa": "AVA",
    "Asignada Vía": "AV",
    "Asignar Vía Nueva": "AVN",
    "Aislada": "AIS",
    "Asignada Unidad Poblacional": "AUP"
}

# Diccionario corregido de municipios con artículo al inicio
municipios_adaptado = {
    "4": "Arafo",
    "5": "Arico",
    "10": "Buenavista del Norte",
    "12": "Fasnia",
    "15": "Garachico",
    "18": "La Guancha",
    "25": "La Matanza de Acentejo",
    "32": "El Rosario",
    "34": "San Juan de la Rambla",
    "39": "Santa Úrsula",
    "40": "Santiago del Teide",
    "41": "El Sauzal",
    "42": "Los Silos",
    "44": "El Tanque",
    "46": "Tegueste",
    "51": "La Victoria de Acentejo",
    "52": "Vilaflor de Chasna"
}

# Mapear nombres de municipios
df['Municipio'] = df['cmun_ine'].astype(str).map(municipios_adaptado)

# === Crear resumen por estado ===
resumen_estado = df['a3_estado'].value_counts().reset_index()
resumen_estado.columns = ['Estado', 'Cantidad']
resumen_estado['Estado'] = resumen_estado['Estado'].map(estado_labels).fillna(resumen_estado['Estado'])
resumen_estado['Porcentaje'] = (resumen_estado['Cantidad'] / resumen_estado['Cantidad'].sum()) * 100
resumen_estado = resumen_estado.sort_values(by="Cantidad", ascending=False)

# === Crear tabla de distribución por municipio y estado ===
tabla_municipio_estado = pd.pivot_table(
    df,
    index='Municipio',
    columns='a3_estado',
    aggfunc='size',
    fill_value=0
)
tabla_municipio_estado = tabla_municipio_estado.rename(columns=estado_labels)

# Añadir columna Total sumando todas las asignaciones por municipio
tabla_municipio_estado['Total'] = tabla_municipio_estado.sum(axis=1)

# Convertir índice Municipio en columna
tabla_municipio_estado = tabla_municipio_estado.reset_index()

# Añadir columna Código al inicio
codigo_municipio = df.groupby('Municipio')['cmun_ine'].first()
tabla_municipio_estado.insert(0, 'Código', tabla_municipio_estado['Municipio'].map(codigo_municipio))

# Reordenar columnas
column_order = [
    'Código',
    'Municipio',
    'Asignada Vía',
    'Asignada Vía Alternativa',
    'Asignada Unidad Poblacional',
    'Asignar Vía Nueva',
    'Aislada',
    'Total'
]
for col in column_order:
    if col not in tabla_municipio_estado.columns:
        tabla_municipio_estado[col] = 0
tabla_municipio_estado = tabla_municipio_estado[column_order]

# === Formatear las columnas de estados con "Cantidad (Porcentaje%)" ===
estado_cols = column_order[2:-1]  # columnas de estados
for col in estado_cols:
    tabla_municipio_estado[col] = tabla_municipio_estado.apply(
        lambda row: f"{row[col]} ({row[col]/row['Total']*100:.1f}%)" if row['Total'] > 0 else "0 (0.0%)",
        axis=1
    )

# === Añadir fila de totales al final ===
total_dict = {'Código':'', 'Municipio':'Total'}
total_general = tabla_municipio_estado['Total'].sum()
for col in estado_cols:
    total_cantidad = tabla_municipio_estado[col].apply(lambda x: int(x.split(' ')[0])).sum()
    porcentaje = total_cantidad / total_general * 100 if total_general > 0 else 0
    total_dict[col] = f"{total_cantidad} ({porcentaje:.1f}%)"
total_dict['Total'] = total_general

fila_total_df = pd.DataFrame([total_dict])
tabla_municipio_estado = pd.concat([tabla_municipio_estado, fila_total_df], ignore_index=True)

# === Gráfico de barras horizontal por estado (ajustado) ===
y_pos = np.arange(len(resumen_estado)) * 0.2  # menor separación vertical
plt.figure(figsize=(4.5,4))  # más estrecho

bars = plt.barh(y_pos, resumen_estado['Cantidad'], height=0.15, align='center', color="#4c72b0")

# Etiquetas del eje Y con mayor tamaño y en negrita
plt.yticks(y_pos, [estado_abrev.get(e, e) for e in resumen_estado['Estado']], fontsize=11, fontweight='bold')

# Eliminamos el título
# plt.title("Distribución por estado de asignación", fontweight='bold')

# Etiquetas de cantidad y porcentaje dentro de las barras con mayor tamaño
for i, (cantidad, porcentaje, estado) in enumerate(zip(resumen_estado['Cantidad'], resumen_estado['Porcentaje'], resumen_estado['Estado'])):
    label = f"{cantidad} ({porcentaje:.1f}%)"
    y = y_pos[i]
    if estado in ["Asignada Vía", "Asignada Vía Alternativa"]:
        plt.text(cantidad * 0.98, y, label, ha='right', va='center',
                 fontsize=11, fontweight='bold', color="white", fontname='DejaVu Sans')
    else:
        plt.text(cantidad + max(resumen_estado['Cantidad']) * 0.01, y, label,
                 va='center', fontsize=11, fontweight='bold', color="dimgray", fontname='DejaVu Sans')

plt.gca().invert_yaxis()
plt.tight_layout()

# Guardar gráfico en SVG y PNG con fondo transparente
plt.savefig("grafico_estado.svg", format="svg", transparent=True)
grafico_path = "grafico_estado.png"
plt.savefig(grafico_path, format="png", transparent=True)
plt.close()


# === Exportar a Excel ===
output_file = "resumen_resultados.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    resumen_estado.to_excel(writer, sheet_name="Por Estado", index=False)
    tabla_municipio_estado.to_excel(writer, sheet_name="Por Municipio", index=False)

# Añadir hoja con gráfico (PNG)
wb = load_workbook(output_file)
ws = wb.create_sheet(title="Gráfico Estado")
img = Image(grafico_path)
img.anchor = 'A1'
ws.add_image(img)
wb.save(output_file)

print(f"Resumen exportado a {output_file} con tabla y gráfico incluidos.")
print(tabla_municipio_estado)
